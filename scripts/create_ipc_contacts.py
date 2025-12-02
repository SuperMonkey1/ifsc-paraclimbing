from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "IPC Athlete Representatives"

# Define headers
headers = ["Name", "Role", "Sport/Organization", "Country"]

# Define the data - extracted from IPC.md
data = [
    # IPC Athletes' Council
    ["Vladyslava Kravchenko", "Chairperson, IPC Athletes' Council (IPC Governing Board Member)", "Para Swimming", "Malta"],
    ["Josh Dueck", "First Vice-Chairperson, IPC Athletes' Council (IPC Governing Board Member)", "Para Alpine Skiing", "Canada"],
    ["Yoomin Won", "Second Vice-Chairperson, IPC Athletes' Council", "Wheelchair Basketball", "South Korea"],
    ["Lenine Cunha", "Council Member (First II Representative)", "Para Athletics", "Portugal"],
    ["Martina Caironi", "Council Member", "Para Athletics", "Italy"],
    ["Denise Schindler", "Council Member", "Para Cycling", "Germany"],
    ["Birgit Skarstein", "Council Member (Dual-Sport Athlete)", "Para Nordic Skiing / Para Rowing", "Norway"],
    ["Yujiao Tan", "Council Member", "Para Powerlifting", "China"],
    ["Husnah Kukundakwe", "Appointed Council Member (African & Youth Rep)", "Para Swimming", "Uganda"],
    ["Grigorios Polychronidis", "Appointed Council Member (High Support Needs Rep)", "Boccia", "Greece"],
    
    # Para Archery
    ["Pippa Britton", "Chair, Para Archery Committee", "Para Archery (World Archery)", "Great Britain"],
    ["Eric Bennett", "Athlete Representative, Para Archery Committee", "Para Archery (World Archery)", "USA"],
    ["Ella Gibson", "Chair, World Archery Athletes' Committee", "Archery (World Archery)", "Great Britain"],
    
    # Para Badminton
    ["Daniel Chan", "Chair, Para Badminton Athletes' Commission (BWF Council Member)", "Para Badminton (BWF)", "Hong Kong"],
    ["Cathrine Rosengren", "Deputy Chair, Para Badminton Athletes' Commission", "Para Badminton (BWF)", "Denmark"],
    ["Amy Burnett", "Commission Member (WH1 Class)", "Para Badminton (BWF)", "USA"],
    ["Guillaume Gailly", "Commission Member (SL4 Class)", "Para Badminton (BWF)", "France"],
    ["Abu Hubaida", "Commission Member (WH2 Class)", "Para Badminton (BWF)", "India"],
    ["Tarek Abbas Gharib Zahry", "Commission Member (WH1 Class)", "Para Badminton (BWF)", "Egypt"],
    
    # Para Canoe
    ["Emma Wiggs", "Female Paracanoe Representative, ICF Athletes' Committee", "Para Canoe (ICF)", "Great Britain"],
    ["Eslam Jahedi", "Male Paracanoe Representative, ICF Athletes' Committee", "Para Canoe (ICF)", "Iran"],
    
    # Para Cycling
    ["Suzanna Tangen", "Female Representative, UCI Athletes' Commission", "Para Cycling (UCI)", "Norway"],
    ["Tristan Bangma", "Male Representative, UCI Athletes' Commission", "Para Cycling (UCI)", "Netherlands"],
    ["Maja Włoszczowska", "Chair, UCI Athletes' Commission (IOC Member)", "Cycling (UCI)", "Poland"],
    
    # Para Equestrian
    ["Erin Orford", "Athlete Representative, FEI Athletes' Committee", "Para Equestrian (FEI)", "Great Britain"],
    
    # Para Rowing
    ["Benjamin Pritchard", "Para Rowing Representative, World Rowing Athletes' Commission", "Para Rowing (World Rowing)", "Great Britain"],
    
    # Para Table Tennis
    ["Kelly van Zon", "Wheelchair Representative, ITTF Athletes' Commission", "Para Table Tennis (ITTF)", "Netherlands"],
    ["Ingela Lundback", "Standing Representative, ITTF Athletes' Commission", "Para Table Tennis (ITTF)", "Sweden"],
    ["Sharath Kamal Achanta", "Co-Chair, ITTF Athletes' Commission", "Table Tennis (ITTF)", "India"],
    ["Liu Shiwen", "Co-Chair, ITTF Athletes' Commission", "Table Tennis (ITTF)", "China"],
    
    # Para Taekwondo
    ["Mehmet Kani Polat", "Chair, World Taekwondo Athlete Committee", "Para Taekwondo (World Taekwondo)", "Turkey"],
    
    # Para Triathlon
    ["Martin Schulz", "Male Representative, World Triathlon Athletes' Committee", "Para Triathlon (World Triathlon)", "Germany"],
    ["Gulnaz Zhuzbaeva", "Female Representative, World Triathlon Athletes' Committee", "Para Triathlon (World Triathlon)", "Kyrgyzstan"],
    
    # Wheelchair Tennis
    ["Lucy Shuker", "Chair, ITF Wheelchair Tennis Player Council", "Wheelchair Tennis (ITF)", "Great Britain"],
    ["Sam Schroder", "Council Member (Quad Division)", "Wheelchair Tennis (ITF)", "Netherlands"],
    ["Nicholas Taylor", "Council Member (Former Player Rep)", "Wheelchair Tennis (ITF)", "USA"],
    ["Charlotte Fairbank", "Council Member", "Wheelchair Tennis (ITF)", "France"],
    ["Najwa Awane", "Council Member (African Rep)", "Wheelchair Tennis (ITF)", "Morocco"],
    
    # World Para Sports (IPC Managed)
    # Para Athletics
    ["Vanessa Low", "Chairperson, Athletes' Advisory Group (Sport Committee & WPA Board Member)", "Para Athletics (World Para Athletics)", "Australia"],
    ["Francisco Cedeno", "Member, Athletes' Advisory Group", "Para Athletics (World Para Athletics)", "Panama"],
    ["Severin Kansa", "Member, Athletes' Advisory Group", "Para Athletics (World Para Athletics)", "Togo"],
    
    # Para Powerlifting
    ["Sherif Osman", "Chairperson, Athlete Advisory Group", "Para Powerlifting (World Para Powerlifting)", "Egypt"],
    ["Marcia Cristina Menezes", "Member, Athlete Advisory Group", "Para Powerlifting (World Para Powerlifting)", "Brazil"],
    ["Maria Markou", "Member, Athlete Advisory Group", "Para Powerlifting (World Para Powerlifting)", "Cyprus"],
    ["Mutaz Al Juneidi", "Member, Athlete Advisory Group", "Para Powerlifting (World Para Powerlifting)", "Jordan"],
    
    # Para Swimming
    ["Ellen Keane", "Chairperson, Athletes' Advisory Group (Sport Committee & WPS Board Member)", "Para Swimming (World Para Swimming)", "Ireland"],
    ["Tess Routliffe", "Member, Athletes' Advisory Group", "Para Swimming (World Para Swimming)", "Canada"],
    ["David Levecq", "Member, Athletes' Advisory Group", "Para Swimming (World Para Swimming)", "Spain"],
    ["Keiichi Kimura", "Member, Athletes' Advisory Group", "Para Swimming (World Para Swimming)", "Japan"],
    
    # Shooting Para Sport
    ["Wanda Jewell", "Chairperson, Athlete Advisory Group", "Shooting Para Sport (World Shooting Para Sport)", "USA"],
    ["Ferrol van Hoeven", "Vice-Chair, Athlete Advisory Group", "Shooting Para Sport (World Shooting Para Sport)", "Netherlands"],
    
    # Para Ice Hockey
    ["Dominic Cozzolino", "Chairperson, Athletes' Committee", "Para Ice Hockey (World Para Ice Hockey)", "Canada"],
    
    # IBSA Sports
    ["Sela Adikinyi Odhiambo", "Chair, IBSA Athletes' Council", "IBSA (Blind Sports)", "Kenya"],
    ["Ricardo Alves", "Male Athlete Representative", "Blind Football (IBSA)", "Brazil"],
    ["Bettina Sulyok", "Female Athlete Representative", "Blind Football (IBSA)", "Austria"],
    ["Alexandre Almeida", "Male Representative", "Goalball (IBSA)", "Portugal"],
    ["Vasiliki Ago", "Female Representative", "Goalball (IBSA)", "Greece"],
    
    # World Abilitysport - Wheelchair Fencing
    ["Yu Chui Yee", "Chair, Wheelchair Fencing Athletes' Council", "Wheelchair Fencing (World Abilitysport)", "Hong Kong"],
    
    # Boccia
    ["Eileen Bartlett", "Chair, Athletes' Committee", "Boccia (BISFed/World Boccia)", "Great Britain"],
    
    # Wheelchair Rugby
    ["Shae Graham", "Board Representative", "Wheelchair Rugby (World Wheelchair Rugby)", "Australia"],
    ["Byron Green", "Athlete Council Member", "Wheelchair Rugby (World Wheelchair Rugby)", "Unknown"],
    ["Trevor Murao", "Athlete Council Member", "Wheelchair Rugby (World Wheelchair Rugby)", "Unknown"],
    
    # Sitting Volleyball
    ["Heidi Peters", "Team Leader/Chair (Board of Directors Member)", "Sitting Volleyball (World ParaVolley)", "Canada"],
    ["Ivan Cosic", "Commission Member", "Sitting Volleyball (World ParaVolley)", "Croatia"],
    ["Masoumeh Barouti", "Commission Member", "Sitting Volleyball (World ParaVolley)", "Iran"],
    
    # Wheelchair Basketball
    ["Erica Gavel", "Chair, Players' Commission (Executive Council Member)", "Wheelchair Basketball (IWBF)", "Canada"],
    ["Alhassan Sedky", "Vice-Chair, Players' Commission", "Wheelchair Basketball (IWBF)", "Egypt"],
    ["Ella Sabljak", "Commission Member", "Wheelchair Basketball (IWBF)", "Australia"],
    ["Terry Bywater", "Commission Member", "Wheelchair Basketball (IWBF)", "Great Britain"],
    
    # Winter Sports - FIS
    ["Adam Hall", "FIS Athletes' Commission & FIS Council Member", "Para Snow Sports (FIS)", "New Zealand"],
    ["Danielle Aravich", "FIS Athletes' Commission Member", "Para Snow Sports (FIS)", "USA"],
    
    # Para Biathlon - IBU
    ["Mark Arendz", "Athlete Representative", "Para Biathlon (IBU)", "Canada"],
    ["Florian Michelon", "Athlete Representative", "Para Biathlon (IBU)", "France"],
    ["Anja Wicker", "Athlete Representative", "Para Biathlon (IBU)", "Germany"],
    
    # Wheelchair Curling
    ["Polina Rozkova", "Wheelchair Curling Representative", "Wheelchair Curling (World Curling)", "Latvia"],
    ["Batoyun Uranchimeg", "Wheelchair Curling Representative", "Wheelchair Curling (World Curling)", "USA"],
    ["Tyler George", "Chair, World Curling Athlete Commission", "Curling (World Curling)", "USA"],
]

# Style definitions
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

# Adjust column widths
column_widths = [25, 60, 45, 18]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze the header row
ws.freeze_panes = "A2"

# Add autofilter
ws.auto_filter.ref = f"A1:D{len(data) + 1}"

# Save the workbook
output_path = r"c:\PYTHON\ifsc-paraclimbing\reference\IPC_Athlete_Representatives.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
print(f"Total contacts: {len(data)}")
